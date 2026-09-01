"""
Chequea, para los ítems con adjunto sospechosamente chico, si el campo
Resumen (abstract) YA arranca con el marcador "[ADJUNTO NO FUNCIONA]" —
o sea, si ya se avisó al autor/a en la ficha del ítem.

Punto de partida: reportes/entradas_a_revisar.xlsx, YA actualizado por
chequear_adjuntos.py (este script necesita correr DESPUÉS, en el mismo
workflow, para que la columna "Bytes" esté fresca — no vuelve a calcular
tamaños, solo lee lo que dejó el chequeo anterior).

De ese Excel toma las filas con Bytes < UMBRAL_BYTES_SOSPECHOSO (150 por
defecto — más ancho que los 42 bytes que usa chequear_adjuntos.py para
marcar "roto", así entran también adjuntos "casi vacíos" que hoy figuran
como OK pero convendría mirar a mano). Para cada una, resuelve el ítem
contra la API de RDU (reutilizando el login y la resolución de link/handle/
uuid/workflowitem de chequear_adjuntos.py) y lee su Resumen.

No modifica nada en RDU ni en el Excel: es de solo lectura. Escribe
reportes/resumen_pendiente.csv con columnas Link, Titulo, Bytes,
TieneMarcador ("OK" si el Resumen YA contiene la frase "ADJUNTO NO
FUNCIONA" en cualquier parte —sin importar corchetes/asteriscos u otro
texto alrededor, porque en la práctica se pegó a mano y no siempre quedó
idéntico al texto de la columna Estado del Excel—, "NO" si todavía no
aparece), ordenado con lo pendiente (NO) primero y, dentro de cada grupo,
por Bytes ascendente.

Variables de entorno opcionales:
  UMBRAL_BYTES_SOSPECHOSO  (default 150)
  CAMPO_RESUMEN            (default "dc.description.abstract" — el nombre
                             estándar de DSpace 7 para el Resumen/abstract;
                             si en RDU se usa otro campo, ajustar acá)
"""
import csv
import os
import re
import sys
import time

import openpyxl

# Reutilizamos login, resolución de links y helpers de chequear_adjuntos.py
# en vez de duplicarlos (viven en el mismo repo, en scripts/).
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from chequear_adjuntos import (  # noqa: E402
    ARCHIVO_ENTRADAS,
    PAUSA_SEGUNDOS,
    _extraer_identificador,
    _login_rdu,
    _metadata,
    _resolver_item,
    _resolver_workflowitem,
)

DIR_REPORTES = "reportes"
ARCHIVO_SALIDA = os.path.join(DIR_REPORTES, "resumen_pendiente.csv")

UMBRAL_BYTES_SOSPECHOSO = int(os.environ.get("UMBRAL_BYTES_SOSPECHOSO", "150"))
CAMPO_RESUMEN = os.environ.get("CAMPO_RESUMEN", "dc.description.abstract")

# Chequeo FLEXIBLE del marcador: en la práctica se pegó a mano y no siempre
# quedó como el texto exacto "[ADJUNTO NO FUNCIONA]" que usa la columna
# Estado del Excel (ej. se encontró un caso real con "[*ADJUNTO NO
# FUNCIONA]"). Por eso acá basta con que la frase "ADJUNTO NO FUNCIONA"
# aparezca en cualquier parte del Resumen, sin importar qué corchetes,
# asteriscos u otros caracteres la rodeen.
_RE_MARCA = re.compile(r"ADJUNTO\s+NO\s+FUNCIONA", re.IGNORECASE)

CAMPOS_SALIDA = ["Link", "Titulo", "Bytes", "TieneMarcador"]


# ---------- LEER LAS FILAS A REVISAR DESDE EL EXCEL YA ACTUALIZADO ----------
def _filas_a_revisar():
    if not os.path.exists(ARCHIVO_ENTRADAS):
        print(f"[ERROR] No existe {ARCHIVO_ENTRADAS}; corré chequear_adjuntos.py primero.")
        return []

    wb = openpyxl.load_workbook(ARCHIVO_ENTRADAS, read_only=True, data_only=True)
    ws = wb.active
    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {nombre: i for i, nombre in enumerate(encabezados)}

    faltan = [c for c in ("Link", "Titulo", "Bytes") if c not in idx]
    if faltan:
        print(f"[ERROR] Al Excel le faltan columnas esperadas: {faltan}")
        return []

    filas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        link = str(row[idx["Link"]] or "").strip()
        if not link:
            continue
        try:
            tam = int(row[idx["Bytes"]])
        except (TypeError, ValueError):
            # Sin tamaño confiable (p.ej. filas en ERROR): no se puede
            # filtrar por bytes, se omite.
            continue
        if tam < UMBRAL_BYTES_SOSPECHOSO:
            filas.append((link, row[idx["Titulo"]] or "", tam))
    return filas


# ---------- RESOLVER CADA LINK CONTRA LA API DE RDU ----------
def _resolver(link: str, opener_autenticado):
    """Devuelve (item o None, opener_autenticado actualizado)."""
    identificador = _extraer_identificador(link)
    if not identificador:
        print(f"  [ERROR] No se pudo interpretar el link '{link}'.")
        return None, opener_autenticado

    tipo, valor = identificador
    try:
        if tipo == "workflowitem":
            if opener_autenticado is None:
                opener_autenticado = _login_rdu() or False
            if not opener_autenticado:
                print(f"  [ERROR] '{link}' es un ítem en revisión y no hay sesión válida "
                      f"(revisá RDU_USER/RDU_PASS).")
                return None, opener_autenticado
            item = _resolver_workflowitem(valor, opener_autenticado)
        else:
            item = _resolver_item(tipo, valor)
    except Exception as e:
        print(f"  [ERROR] Falló la consulta a RDU para '{link}': {e}")
        return None, opener_autenticado

    if not item:
        print(f"  [ERROR] No se encontró el ítem para '{link}'.")
    return item, opener_autenticado


def main():
    filas = _filas_a_revisar()
    print(f"[RESUMEN] {len(filas)} entrada(s) con Bytes < {UMBRAL_BYTES_SOSPECHOSO} para revisar "
          f"(campo Resumen: {CAMPO_RESUMEN}).")

    opener_autenticado = None
    salida = []
    errores = 0

    for link, titulo_excel, tam in filas:
        item, opener_autenticado = _resolver(link, opener_autenticado)
        if not item:
            errores += 1
            continue

        resumen = _metadata(item, CAMPO_RESUMEN)
        tiene_marcador = "OK" if _RE_MARCA.search(resumen) else "NO"
        titulo = item.get("name") or titulo_excel

        salida.append({"Link": link, "Titulo": titulo, "Bytes": tam, "TieneMarcador": tiene_marcador})
        print(f"  {tiene_marcador} — {tam}b — {titulo}")

        time.sleep(PAUSA_SEGUNDOS)

    # Lo pendiente (NO) primero; dentro de cada grupo, por Bytes ascendente.
    salida.sort(key=lambda f: (f["TieneMarcador"] == "OK", f["Bytes"]))

    os.makedirs(DIR_REPORTES, exist_ok=True)
    with open(ARCHIVO_SALIDA, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CAMPOS_SALIDA)
        writer.writeheader()
        writer.writerows(salida)

    faltan = sum(1 for f in salida if f["TieneMarcador"] == "NO")
    print(f"[RESUMEN] {len(salida)} entrada(s) escritas en {ARCHIVO_SALIDA}: "
          f"{faltan} todavía SIN marcador, {len(salida) - faltan} ya OK. {errores} error(es).")

    if errores:
        sys.exit(1)


if __name__ == "__main__":
    main()
