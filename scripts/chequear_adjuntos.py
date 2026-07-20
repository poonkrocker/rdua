"""
Chequeo de adjuntos rotos en RDU a partir de una lista de Excel.

Vos controlás EXACTAMENTE qué ítems se revisan: pegás el link (o el handle)
de cada uno en la columna "Link" de reportes/entradas_a_revisar.xlsx, y el
script solo chequea esas filas — nada de recorrer todo el repositorio ni
filtrar por tipo/año contra el buscador. Por cada fila, deja el resultado
(Estado / Detalle / Bytes / ÚltimoChequeo) en las columnas de al lado, en el
mismo archivo. No modifica nada en RDU: es de solo lectura.

La columna "Link" acepta cualquier forma en la que lo hayas copiado:
  - link completo con handle:      https://rdu.unc.edu.ar/handle/11086/29993
  - link completo con uuid:        https://rdu.unc.edu.ar/items/<uuid>
  - el handle pelado:              11086/29993
  - el uuid pelado
  - link de edición de un ítem TODAVÍA EN REVISIÓN (no público):
    https://rdu.unc.edu.ar/workflowitems/<id>/edit — este caso necesita
    RDU_USER/RDU_PASS configurados (los mismos secrets de "Procesar items
    RDU"), porque un ítem en revisión no es visible sin login.

Si el archivo de entradas no existe todavía, el script lo crea vacío (con
los encabezados) y termina — pegá los links y volvé a correrlo.

Además, aparte del Excel, si hay RDU_USER/RDU_PASS configurados, el script
también intenta listar TODOS los ítems en revisión (sin que hayas pegado
sus links a mano) y deja ese resultado en
reportes/adjuntos_rotos_en_revision.csv. Es best-effort: en la práctica esa
lista completa puede devolver 403 si la cuenta no tiene rol de admin —  en
ese caso, revisar por link individual (arriba) suele funcionar igual porque
es un permiso distinto. Cualquier falla ahí se avisa por log sin afectar el
chequeo del Excel.
"""
import csv
import http.cookiejar
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date

import openpyxl

BASE_URL = os.environ.get("RDU_BASE_URL", "https://rdu.unc.edu.ar").rstrip("/")
API = f"{BASE_URL}/server/api"

UMBRAL_BYTES_VACIO = int(os.environ.get("UMBRAL_BYTES_VACIO", "42"))
PAUSA_SEGUNDOS = float(os.environ.get("PAUSA_SEGUNDOS", "0.3"))  # para no saturar el servidor
REINTENTOS = int(os.environ.get("REINTENTOS", "3"))

DIR_REPORTES = "reportes"
ARCHIVO_ENTRADAS = os.environ.get("ARCHIVO_ENTRADAS", os.path.join(DIR_REPORTES, "entradas_a_revisar.xlsx"))
ARCHIVO_REPORTE_WORKFLOW = os.path.join(DIR_REPORTES, "adjuntos_rotos_en_revision.csv")

MARCA = "[ADJUNTO NO FUNCIONA]"

COLUMNAS_ENTRADAS = ["Link", "Titulo", "Estado", "Detalle", "Bytes", "UltimoChequeo"]
CAMPOS_REPORTE_WORKFLOW = ["handle", "titulo", "tipo", "problema", "bytes", "link", "detectado"]

# Algunos WAF/firewalls institucionales cortan la conexión sin responder
# cuando ven el user-agent por defecto de Python ("Python-urllib/3.x"),
# porque lo reconocen como firma de bot. Mandamos cabeceras de navegador
# real para evitarlo.
CABECERAS_BASE = {
    "Accept": "application/json",
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Accept-Language": "es-AR,es;q=0.9,en;q=0.8",
}

_RE_UUID = re.compile(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", re.IGNORECASE)
_RE_HANDLE = re.compile(r"\b(\d+/\d+)\b")


# ---------- HTTP CON REINTENTOS ----------
def _abrir_con_reintentos(opener, req):
    ultimo_error = None
    for intento in range(1, REINTENTOS + 1):
        try:
            return opener.open(req, timeout=40)
        except urllib.error.HTTPError as e:
            # Los 4xx (403 sin permiso, 404 no existe, etc.) no son
            # problemas de red pasajeros: reintentar no los va a arreglar.
            if 400 <= e.code < 500:
                raise
            ultimo_error = e
            if intento < REINTENTOS:
                print(f"  [WARN] Falló la petición (intento {intento}/{REINTENTOS}): {e}. Reintentando...")
                time.sleep(1.5 * intento)
        except Exception as e:
            ultimo_error = e
            if intento < REINTENTOS:
                print(f"  [WARN] Falló la petición (intento {intento}/{REINTENTOS}): {e}. Reintentando...")
                time.sleep(1.5 * intento)
    raise ultimo_error


_OPENER_ANONIMO = urllib.request.build_opener()


def _get(url: str, opener=None) -> dict:
    opener = opener or _OPENER_ANONIMO
    req = urllib.request.Request(url, headers=CABECERAS_BASE)
    resp = _abrir_con_reintentos(opener, req)
    return json.loads(resp.read())


def _metadata(item: dict, campo: str) -> str:
    valores = item.get("metadata", {}).get(campo, [])
    return "; ".join(v.get("value", "") for v in valores if v.get("value"))


def _evaluar_bundle_original(item: dict):
    """Devuelve (problema, bytes) si el adjunto ORIGINAL está roto, o (None, bytes) si está OK
    (el tamaño devuelto cuando está OK es informativo, el del archivo más grande del bundle)."""
    bundles = item.get("_embedded", {}).get("bundles", {}).get("_embedded", {}).get("bundles", [])
    original = next((b for b in bundles if b.get("name") == "ORIGINAL"), None)
    if not original:
        return "SIN ARCHIVO ADJUNTO", 0

    bitstreams_meta = original.get("_embedded", {}).get("bitstreams", {})
    bitstreams = bitstreams_meta.get("_embedded", {}).get("bitstreams", [])
    if not bitstreams:
        return "SIN ARCHIVO ADJUNTO", 0

    if bitstreams_meta.get("page", {}).get("totalPages", 1) > 1:
        print(f"  [WARN] {item.get('handle')} tiene más de {len(bitstreams)} archivos originales; "
              f"solo se revisaron los primeros {len(bitstreams)}.")

    for b in bitstreams:
        tam = b.get("sizeBytes", 0)
        if tam <= UMBRAL_BYTES_VACIO:
            return "ADJUNTO VACÍO/ROTO", tam

    return None, max(b.get("sizeBytes", 0) for b in bitstreams)


# ---------- RESOLVER UN LINK/HANDLE/UUID PEGADO A MANO A UN ÍTEM DE RDU ----------
def _extraer_identificador(texto: str):
    """A partir de lo que se pegó en la columna Link, devuelve ("uuid", valor),
    ("handle", valor) o ("workflowitem", id). Acepta link completo (con
    /items/<uuid>, /handle/<prefijo>/<sufijo> o /workflowitems/<id>/edit —
    este último es un ítem TODAVÍA EN REVISIÓN, no público), o el uuid/handle
    pelados. None si no se pudo interpretar nada."""
    texto = (texto or "").strip()
    if not texto:
        return None

    # Ítem en revisión (no público): el link de edición usa el ID interno
    # del workflowitem, NO el uuid del ítem — hace falta resolverlo vía la
    # API autenticada (ver _resolver_workflowitem).
    m = re.search(r"/workflowitems/(\d+)", texto)
    if m:
        return ("workflowitem", m.group(1))

    m = re.search(r"/handle/(\d+/\d+)", texto)
    if m:
        return ("handle", m.group(1))

    m_uuid = _RE_UUID.search(texto)
    if m_uuid and ("/items/" in texto.lower() or texto.lower() == m_uuid.group(0).lower()):
        return ("uuid", m_uuid.group(0))

    m = _RE_HANDLE.search(texto)
    if m:
        return ("handle", m.group(1))

    if m_uuid:
        return ("uuid", m_uuid.group(0))

    return None


def _resolver_item(tipo: str, valor: str) -> dict | None:
    if tipo == "handle":
        data = _get(f"{API}/pid/find?id={urllib.parse.quote(valor, safe='')}")
        uuid = data.get("uuid") if data else None
        if not uuid:
            return None
    else:
        uuid = valor
    return _get(f"{API}/core/items/{uuid}?embed=bundles/bitstreams")


def _resolver_workflowitem(workflowitem_id: str, opener) -> dict | None:
    """Resuelve un ítem TODAVÍA EN REVISIÓN a partir del id de su
    workflowitem (el que aparece en el link .../workflowitems/<id>/edit).
    Requiere una sesión autenticada (ver _login_rdu) — a diferencia de
    listar TODOS los workflowitems (que devolvió 403, probablemente por
    requerir un rol de admin), pedir uno puntual por id debería andar con
    cualquier cuenta que tenga la tarea asignada. No confirmado contra un
    caso real; si también da 403, avisá para ajustar el enfoque."""
    data = _get(
        f"{API}/workflow/workflowitems/{workflowitem_id}?embed=item&embed=item/bundles/bitstreams",
        opener=opener,
    )
    return data.get("_embedded", {}).get("item")


# ---------- CHEQUEO PRINCIPAL: LOS ÍTEMS QUE PUSISTE EN EL EXCEL ----------
def _crear_planilla_vacia():
    os.makedirs(os.path.dirname(ARCHIVO_ENTRADAS) or ".", exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entradas"
    ws.append(COLUMNAS_ENTRADAS)
    wb.save(ARCHIVO_ENTRADAS)
    print(f"No existía {ARCHIVO_ENTRADAS}; se creó vacío con los encabezados: "
          f"{', '.join(COLUMNAS_ENTRADAS)}.")
    print("Pegá los links (o handles) a revisar en la columna 'Link' y volvé a correr el workflow.")


def _indices_columnas(ws) -> dict:
    """Lee la fila de encabezados; agrega las columnas que falten al final."""
    encabezados = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    indices = {}
    for nombre in COLUMNAS_ENTRADAS:
        if nombre not in encabezados:
            col = len(encabezados) + 1
            ws.cell(row=1, column=col, value=nombre)
            encabezados.append(nombre)
        indices[nombre] = encabezados.index(nombre) + 1
    return indices


def _escribir_resultado_fila(ws, indices: dict, fila: int, estado: str, detalle: str, tam, titulo: str = None):
    if titulo is not None and indices.get("Titulo"):
        ws.cell(row=fila, column=indices["Titulo"], value=titulo)
    ws.cell(row=fila, column=indices["Estado"], value=estado)
    ws.cell(row=fila, column=indices["Detalle"], value=detalle)
    ws.cell(row=fila, column=indices["Bytes"], value=tam)
    ws.cell(row=fila, column=indices["UltimoChequeo"], value=date.today().isoformat())


def chequeo_desde_excel():
    if not os.path.exists(ARCHIVO_ENTRADAS):
        _crear_planilla_vacia()
        return

    wb = openpyxl.load_workbook(ARCHIVO_ENTRADAS)
    ws = wb.active
    indices = _indices_columnas(ws)

    total = 0
    rotos = 0
    ok = 0
    errores = 0

    # Login lazy: solo se intenta si aparece algún link de workflowitem (ítem
    # todavía en revisión, no público). None = todavía no se intentó,
    # False = se intentó y falló (no reintentar de nuevo por cada fila).
    opener_autenticado = None

    for fila in range(2, ws.max_row + 1):
        link = str(ws.cell(row=fila, column=indices["Link"]).value or "").strip()
        if not link:
            continue
        total += 1

        identificador = _extraer_identificador(link)
        if not identificador:
            _escribir_resultado_fila(ws, indices, fila, "ERROR", "No se pudo interpretar el link/handle pegado", "")
            errores += 1
            print(f"  [ERROR] Fila {fila}: no se pudo interpretar '{link}'")
            continue

        tipo, valor = identificador
        try:
            if tipo == "workflowitem":
                if opener_autenticado is None:
                    opener_autenticado = _login_rdu() or False
                if not opener_autenticado:
                    _escribir_resultado_fila(
                        ws, indices, fila, "ERROR",
                        "Es un ítem en revisión (workflowitem): hace falta RDU_USER/RDU_PASS "
                        "configurados y con permiso para verlo", "",
                    )
                    errores += 1
                    print(f"  [ERROR] Fila {fila}: '{link}' es un ítem en revisión y no hay sesión válida.")
                    continue
                item = _resolver_workflowitem(valor, opener_autenticado)
            else:
                item = _resolver_item(tipo, valor)
        except Exception as e:
            _escribir_resultado_fila(ws, indices, fila, "ERROR", f"Falló la consulta a RDU: {e}", "")
            errores += 1
            print(f"  [ERROR] Fila {fila} ({link}): {e}")
            continue

        if not item:
            _escribir_resultado_fila(ws, indices, fila, "ERROR", "No se encontró ese ítem en RDU", "")
            errores += 1
            print(f"  [ERROR] Fila {fila}: no se encontró el ítem para '{link}'")
            continue

        titulo = item.get("name", "")
        problema, tam = _evaluar_bundle_original(item)
        if problema:
            _escribir_resultado_fila(ws, indices, fila, MARCA, problema, tam, titulo=titulo)
            rotos += 1
            print(f"  {MARCA} fila {fila} ({item.get('handle', link)}): {problema} ({tam} bytes) — {titulo}")
        else:
            _escribir_resultado_fila(ws, indices, fila, "OK", "", tam, titulo=titulo)
            ok += 1

        time.sleep(PAUSA_SEGUNDOS)

    wb.save(ARCHIVO_ENTRADAS)
    print(f"[EXCEL] {total} entrada(s) revisada(s): {ok} OK, {rotos} rota(s), {errores} con error. "
          f"Guardado en {ARCHIVO_ENTRADAS}.")


# ---------- CHEQUEO OPCIONAL: ÍTEMS EN REVISIÓN (TODAVÍA NO PÚBLICOS, requiere login) ----------
def _leer_reporte_workflow() -> dict:
    if not os.path.exists(ARCHIVO_REPORTE_WORKFLOW):
        return {}
    with open(ARCHIVO_REPORTE_WORKFLOW, newline="", encoding="utf-8") as f:
        return {fila["handle"]: fila for fila in csv.DictReader(f)}


def _guardar_reporte_workflow(filas: dict):
    os.makedirs(DIR_REPORTES, exist_ok=True)
    with open(ARCHIVO_REPORTE_WORKFLOW, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CAMPOS_REPORTE_WORKFLOW)
        writer.writeheader()
        for handle in sorted(filas):
            writer.writerow(filas[handle])


def _fila_reporte_workflow(item: dict, problema: str, tam: int, handle_alternativo: str = "") -> dict:
    handle = item.get("handle") or handle_alternativo
    return {
        "handle": handle,
        "titulo": item.get("name", ""),
        "tipo": _metadata(item, "dc.type"),
        "problema": f"{MARCA} {problema} (EN REVISIÓN, no público)",
        "bytes": tam,
        "link": f"{BASE_URL}/handle/{handle}" if item.get("handle") else "(sin handle: revisar en el panel de RDU)",
        "detectado": date.today().isoformat(),
    }


def _login_rdu():
    """Autentica contra la API de RDU. Devuelve un opener de urllib listo para
    usar en pedidos autenticados, o None si no hay credenciales o el login
    falló (en cuyo caso el chequeo de workflow simplemente se omite)."""
    usuario = os.environ.get("RDU_USER")
    clave = os.environ.get("RDU_PASS")
    if not usuario or not clave:
        print("[WORKFLOW] RDU_USER/RDU_PASS no están configurados; se omite el chequeo de ítems en revisión.")
        return None

    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))

    req = urllib.request.Request(f"{API}/authn/status", headers=CABECERAS_BASE)
    resp = _abrir_con_reintentos(opener, req)
    xsrf = resp.headers.get("DSPACE-XSRF-TOKEN")
    if not xsrf:
        print("[WORKFLOW] No se pudo obtener el token CSRF de RDU; se omite el chequeo de ítems en revisión.")
        return None

    body = urllib.parse.urlencode({"user": usuario, "password": clave}).encode()
    req = urllib.request.Request(
        f"{API}/authn/login", data=body, method="POST",
        headers={
            **CABECERAS_BASE,
            "Content-Type": "application/x-www-form-urlencoded",
            "X-XSRF-TOKEN": xsrf,
        },
    )
    try:
        resp = _abrir_con_reintentos(opener, req)
    except urllib.error.HTTPError as e:
        print(f"[WORKFLOW] Login a RDU falló ({e.code}): revisá RDU_USER/RDU_PASS. Se omite este chequeo.")
        return None

    token = resp.headers.get("Authorization")
    if not token:
        print("[WORKFLOW] El login a RDU no devolvió token de autorización; se omite este chequeo.")
        return None

    opener.addheaders = [("Authorization", token)] + list(CABECERAS_BASE.items())
    print("[WORKFLOW] Login a RDU exitoso.")
    return opener


def chequeo_workflow():
    """Revisa ítems que todavía están en el circuito de revisión de DSpace
    (no públicos, por eso no tienen un link para poner en el Excel).
    Best-effort: cualquier error (incluido el login o un 403 por falta de
    permisos de revisor/admin en la cuenta) se loguea y no afecta el
    chequeo del Excel."""
    print("[WORKFLOW] Buscando ítems en revisión (no públicos) con adjunto roto...")
    filas_reporte = _leer_reporte_workflow()
    try:
        opener = _login_rdu()
        if not opener:
            return

        encontrados = 0
        pagina = 0
        while True:
            url = (f"{API}/workflow/workflowitems"
                   f"?embed=item&embed=item/bundles/bitstreams&size=20&page={pagina}")
            data = _get(url, opener=opener)
            resultado = data.get("_embedded", {})
            workflowitems = resultado.get("workflowitems", [])

            for wi in workflowitems:
                item = wi.get("_embedded", {}).get("item")
                if not item:
                    continue
                handle = item.get("handle") or f"workflow-{wi.get('id', '')}"
                problema, tam = _evaluar_bundle_original(item)
                if problema:
                    filas_reporte[handle] = _fila_reporte_workflow(item, problema, tam, handle)
                    encontrados += 1
                    print(f"  {MARCA} {handle} — {problema} (en revisión) — {item.get('name', '')}")

            pagina_info = data.get("page", {})
            total_paginas = pagina_info.get("totalPages", 1)
            pagina += 1
            if pagina >= total_paginas:
                break
            time.sleep(PAUSA_SEGUNDOS)

        print(f"[WORKFLOW] {encontrados} ítem(s) en revisión con adjunto roto.")
    except Exception as e:
        print(f"[WORKFLOW] [ERROR] No se pudo completar el chequeo de ítems en revisión: {e}. "
              f"Puede ser un problema de permisos (la cuenta necesita rol de revisor/admin en RDU "
              f"para ver la cola de workflow) o que el endpoint/formato necesite ajuste.")
    finally:
        _guardar_reporte_workflow(filas_reporte)


def main():
    hubo_fallas = False
    try:
        chequeo_desde_excel()
    except Exception as e:
        hubo_fallas = True
        print(f"[ERROR] El chequeo del Excel falló: {e}")

    try:
        chequeo_workflow()
    except Exception as e:
        hubo_fallas = True
        print(f"[ERROR] El chequeo de workflow falló: {e}")

    if hubo_fallas:
        sys.exit(1)


if __name__ == "__main__":
    main()
